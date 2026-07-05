"""Excel export for the scan ranking (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from windvane.scan.pipeline import ScanResult

HEADERS = ["#", "Symbol", "Sector", "Raw", "Critiqued", "z", "percentile", "Narrative"]
COL_WIDTHS = [5, 10, 24, 8, 10, 8, 11, 90]

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")  # green (project theme)
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOP_FILL = PatternFill("solid", fgColor="E8F5E9")  # light green for top-N


def write_excel(
    path: Path,
    rows: list[tuple[ScanResult, float, float]],
    *,
    top: int = 20,
) -> None:
    """Write the ranking to an .xlsx with a styled header and top-N highlight."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Ranking"

    # Header
    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for rank, (r, z, pct) in enumerate(rows, 1):
        ws.append(
            [
                rank,
                r.symbol,
                r.sector or "-",
                round(r.raw_strength, 3),
                round(r.raw_strength_critiqued, 3),
                round(z, 2),
                round(pct, 1),
                r.narrative_summary or "",
            ]
        )
        if rank <= top:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=rank + 1, column=col).fill = TOP_FILL

    # Column widths + wrap narrative
    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row_cells in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze header, add autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
